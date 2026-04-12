import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Color palette ──
DARK_BG = "1B1B2F"
ACCENT = "E94560"
ACCENT2 = "0F3460"
ACCENT3 = "16213E"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
MEDIUM_GRAY = "D9D9D9"
GREEN = "27AE60"
YELLOW = "F39C12"
RED = "E74C3C"
BLUE = "3498DB"
PURPLE = "9B59B6"
ORANGE = "E67E22"
TEAL = "1ABC9C"

# ── Reusable styles ──
title_font = Font(name="Calibri", size=20, bold=True, color=WHITE)
subtitle_font = Font(name="Calibri", size=14, bold=True, color=WHITE)
header_font = Font(name="Calibri", size=11, bold=True, color=WHITE)
body_font = Font(name="Calibri", size=11, color="333333")
body_bold = Font(name="Calibri", size=11, bold=True, color="333333")
small_font = Font(name="Calibri", size=9, color="888888")

title_fill = PatternFill(start_color=ACCENT3, end_color=ACCENT3, fill_type="solid")
header_fill = PatternFill(start_color=ACCENT2, end_color=ACCENT2, fill_type="solid")
accent_fill = PatternFill(start_color=ACCENT, end_color=ACCENT, fill_type="solid")
light_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
white_fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
green_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
yellow_fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin", color=MEDIUM_GRAY),
    right=Side(style="thin", color=MEDIUM_GRAY),
    top=Side(style="thin", color=MEDIUM_GRAY),
    bottom=Side(style="thin", color=MEDIUM_GRAY),
)


def style_header_row(ws, row, cols, fill=None):
    f = fill or header_fill
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = f
        cell.alignment = center
        cell.border = thin_border


def style_data_row(ws, row, cols, alt=False):
    f = light_fill if alt else white_fill
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = body_font
        cell.fill = f
        cell.alignment = center
        cell.border = thin_border


def add_title_banner(ws, text, row, cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1)
    cell.value = text
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 45


def add_subtitle(ws, text, row, cols, fill=None):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1)
    cell.value = text
    cell.font = subtitle_font
    cell.fill = fill or accent_fill
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 30


# ═══════════════════════════════════════════════════════════════
# SHEET 1: RESUMEN EJECUTIVO
# ═══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Resumen Ejecutivo"
ws1.sheet_properties.tabColor = ACCENT

for c in range(1, 8):
    ws1.column_dimensions[get_column_letter(c)].width = 22

add_title_banner(ws1, "INFORME DE COSTOS - BOT DE VENTAS INSTAGRAM", 1, 7)
add_subtitle(ws1, "  Scuderia St4ge - Academia de Kartismo | Abril 2026", 2, 7,
             fill=PatternFill(start_color=ACCENT3, end_color=ACCENT3, fill_type="solid"))

# Key metrics cards
r = 4
ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
ws1.cell(row=r, column=1, value="METRICAS CLAVE").font = Font(name="Calibri", size=13, bold=True, color=ACCENT2)

r = 5
metrics = [
    ("Costo por Venta\n(solo texto)", "$0.13 USD", green_fill),
    ("Costo por Venta\n(con audio)", "$0.35 USD", yellow_fill),
    ("Infraestructura\nbase/mes", "$10 USD", light_fill),
    ("100 leads/mes\ncosto total", "$24 USD", green_fill),
    ("Tasa conversion\nestimada", "15-25%", yellow_fill),
    ("Mensajes por\nventa promedio", "8-15", light_fill),
]

for i, (label, value, fill) in enumerate(metrics):
    col = i + 1
    cell_label = ws1.cell(row=r, column=col, value=label)
    cell_label.font = Font(name="Calibri", size=9, color="666666")
    cell_label.fill = fill
    cell_label.alignment = center
    cell_label.border = thin_border

    cell_val = ws1.cell(row=r + 1, column=col, value=value)
    cell_val.font = Font(name="Calibri", size=16, bold=True, color=ACCENT2)
    cell_val.fill = fill
    cell_val.alignment = center
    cell_val.border = thin_border

ws1.row_dimensions[r].height = 35
ws1.row_dimensions[r + 1].height = 40

# Technology stack
r = 8
add_subtitle(ws1, "  Stack Tecnologico", r, 7, fill=PatternFill(start_color=ACCENT2, end_color=ACCENT2, fill_type="solid"))
r = 9
headers = ["Componente", "Tecnologia", "Funcion", "Costo/mes (base)", "Costo/mes (500 leads)", "Escalable", "Notas"]
for i, h in enumerate(headers):
    ws1.cell(row=r, column=i + 1, value=h)
style_header_row(ws1, r, 7)

stack_data = [
    ["Backend", "Go + Fiber", "Servidor webhook", "$0", "$0", "Si", "Compilado, ultra eficiente"],
    ["IA Conversacional", "Claude Sonnet 4.6", "Respuestas de venta", "$5-10", "$50", "Si", "Pay per use"],
    ["Voz (TTS)", "ElevenLabs v2", "Voz clonada del dueno", "$2-5", "$18", "Si", "Voz personalizada"],
    ["Voz (STT)", "ElevenLabs Scribe", "Transcribir audios", "$1-2", "$8", "Si", "Espanol nativo"],
    ["Base de Datos", "SQLite/PostgreSQL", "Historial, leads", "$0-7", "$7", "Si", "Self-hosted posible"],
    ["Knowledge Base", "Google Sheets", "Info del negocio", "$0", "$0", "Si", "Editable por el dueno"],
    ["Hosting", "VPS (Hetzner/DO)", "Servidor", "$5-12", "$15", "Si", "Solo si crece mucho"],
]

for i, row_data in enumerate(stack_data):
    r_idx = r + 1 + i
    for j, val in enumerate(row_data):
        ws1.cell(row=r_idx, column=j + 1, value=val)
    style_data_row(ws1, r_idx, 7, alt=i % 2 == 0)

# How it works
r = 18
add_subtitle(ws1, "  Como Funciona el Bot", r, 7, fill=PatternFill(start_color=ACCENT2, end_color=ACCENT2, fill_type="solid"))

flow_data = [
    ["PASO", "Accion", "Detalle", "Costo", "", "", ""],
    ["1", "Cliente escribe por Instagram DM", "Mensaje de texto o nota de voz", "Gratis", "", "", ""],
    ["2", "Bot detecta intencion (local)", "Saludo, precio, horario, compra, objecion...", "$0.00", "", "", ""],
    ["3", "Se calcula Lead Score", "Sistema de puntos: +5/msg, +20 precio, +30 compra", "$0.00", "", "", ""],
    ["4", "Se selecciona estrategia de venta", "Informar, persuadir, guiar, cerrar, manejar objecion", "$0.00", "", "", ""],
    ["5", "Claude genera respuesta natural", "Tono colombiano, max 2 oraciones, con pregunta", "~$0.011", "", "", ""],
    ["6", "Si es audio: voz clonada responde", "ElevenLabs con la voz real del dueno", "~$0.02-0.04", "", "", ""],
    ["7", "Bot envia respuesta al cliente", "Texto o audio segun como escribio el cliente", "Gratis", "", "", ""],
]

for i, row_data in enumerate(flow_data):
    r_idx = r + 1 + i
    for j, val in enumerate(row_data):
        ws1.cell(row=r_idx, column=j + 1, value=val)
    if i == 0:
        style_header_row(ws1, r_idx, 7)
    else:
        style_data_row(ws1, r_idx, 7, alt=i % 2 == 0)
        ws1.cell(row=r_idx, column=1).font = Font(name="Calibri", size=14, bold=True, color=ACCENT)

# ═══════════════════════════════════════════════════════════════
# SHEET 2: COSTOS POR CONVERSACION
# ═══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Costos por Conversacion")
ws2.sheet_properties.tabColor = GREEN

for c in range(1, 7):
    ws2.column_dimensions[get_column_letter(c)].width = 22

add_title_banner(ws2, "DESGLOSE DE COSTOS POR CONVERSACION", 1, 6)

r = 3
add_subtitle(ws2, "  Costo por Llamada de API", r, 6, fill=PatternFill(start_color=ACCENT2, end_color=ACCENT2, fill_type="solid"))
r = 4
headers = ["Componente", "Cuando se usa", "Costo por uso", "Tokens/Caracteres", "Modelo", "Escalabilidad"]
for i, h in enumerate(headers):
    ws2.cell(row=r, column=i + 1, value=h)
style_header_row(ws2, r, 6)

api_data = [
    ["Intent Detection", "Cada mensaje", "$0.000", "N/A (local)", "Keywords Go", "Infinita"],
    ["Lead Scoring", "Cada mensaje", "$0.000", "N/A (local)", "Algoritmo Go", "Infinita"],
    ["Claude Sonnet 4.6 (Input)", "Cada respuesta", "$0.0084", "~2,800 tokens", "$3/MTok", "Linear"],
    ["Claude Sonnet 4.6 (Output)", "Cada respuesta", "$0.0030", "~200 tokens", "$15/MTok", "Linear"],
    ["ElevenLabs STT", "Solo audio entrante", "$0.010/min", "~30s promedio", "Scribe v1", "Linear"],
    ["ElevenLabs TTS", "Solo audio saliente", "$0.018/1K chars", "~200 chars/resp", "Multilingual v2", "Linear"],
]

for i, row_data in enumerate(api_data):
    r_idx = r + 1 + i
    for j, val in enumerate(row_data):
        ws2.cell(row=r_idx, column=j + 1, value=val)
    style_data_row(ws2, r_idx, 6, alt=i % 2 == 0)

# Cost per conversation type
r = 12
add_subtitle(ws2, "  Costo por Tipo de Conversacion", r, 6, fill=PatternFill(start_color=ACCENT2, end_color=ACCENT2, fill_type="solid"))
r = 13
headers = ["Tipo de Conversacion", "Mensajes", "Costo Claude", "Costo Voz", "Costo Total", "Resultado"]
for i, h in enumerate(headers):
    ws2.cell(row=r, column=i + 1, value=h)
style_header_row(ws2, r, 6)

conv_data = [
    ["Venta rapida (texto)", 8, 0.088, 0, 0.09, "VENTA"],
    ["Venta media (texto)", 12, 0.132, 0, 0.13, "VENTA"],
    ["Venta larga (texto)", 20, 0.220, 0, 0.22, "VENTA"],
    ["Venta rapida (audio)", 8, 0.088, 0.15, 0.24, "VENTA"],
    ["Venta media (audio)", 12, 0.132, 0.22, 0.35, "VENTA"],
    ["Venta larga (audio)", 20, 0.220, 0.36, 0.58, "VENTA"],
    ["Consulta sin venta", 5, 0.055, 0, 0.06, "NO VENTA"],
    ["Lead frio (abandona)", 3, 0.033, 0, 0.03, "NO VENTA"],
]

for i, row_data in enumerate(conv_data):
    r_idx = r + 1 + i
    ws2.cell(row=r_idx, column=1, value=row_data[0])
    ws2.cell(row=r_idx, column=2, value=row_data[1])
    for j in range(2, 5):
        cell = ws2.cell(row=r_idx, column=j + 1, value=row_data[j])
        cell.number_format = '$#,##0.000'
    ws2.cell(row=r_idx, column=6, value=row_data[5])
    style_data_row(ws2, r_idx, 6, alt=i % 2 == 0)

    result_cell = ws2.cell(row=r_idx, column=6)
    if row_data[5] == "VENTA":
        result_cell.font = Font(name="Calibri", size=11, bold=True, color=GREEN)
    else:
        result_cell.font = Font(name="Calibri", size=11, bold=True, color=RED)

# Chart: Cost per conversation type
chart1 = BarChart()
chart1.type = "col"
chart1.title = "Costo por Tipo de Conversacion (USD)"
chart1.y_axis.title = "Costo (USD)"
chart1.x_axis.title = "Tipo"
chart1.style = 10
chart1.width = 28
chart1.height = 16

data_ref = Reference(ws2, min_col=3, max_col=5, min_row=13, max_row=21)
cats = Reference(ws2, min_col=1, min_row=14, max_row=21)
chart1.add_data(data_ref, titles_from_data=True)
chart1.set_categories(cats)
chart1.shape = 4

chart1.series[0].graphicalProperties.solidFill = BLUE
chart1.series[1].graphicalProperties.solidFill = ORANGE
chart1.series[2].graphicalProperties.solidFill = GREEN

ws2.add_chart(chart1, "A23")

# ═══════════════════════════════════════════════════════════════
# SHEET 3: PROYECCION POR VOLUMEN
# ═══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Proyeccion Mensual")
ws3.sheet_properties.tabColor = BLUE

for c in range(1, 8):
    ws3.column_dimensions[get_column_letter(c)].width = 20

add_title_banner(ws3, "PROYECCION DE COSTOS MENSUALES POR VOLUMEN", 1, 7)

r = 3
add_subtitle(ws3, "  Supuestos del Modelo", r, 7, fill=PatternFill(start_color=ACCENT2, end_color=ACCENT2, fill_type="solid"))

assumptions = [
    ["Tasa de conversion", "15-25%", "Porcentaje de leads que compran"],
    ["Distribucion canales", "70% texto / 20% mixto / 10% audio", "Como interactuan los leads"],
    ["Msgs por venta (texto)", "12 promedio", "Conversacion tipica hasta cierre"],
    ["Msgs por consulta sin venta", "5 promedio", "Lead que no convierte"],
    ["Costo Claude por msg", "$0.011 USD", "Input + output tokens promedio"],
    ["Costo ElevenLabs por audio", "$0.03 USD", "STT + TTS promedio por mensaje"],
]

r = 4
headers = ["Parametro", "Valor", "Descripcion", "", "", "", ""]
for i, h in enumerate(headers):
    ws3.cell(row=r, column=i + 1, value=h)
style_header_row(ws3, r, 7)

for i, row_data in enumerate(assumptions):
    r_idx = r + 1 + i
    for j, val in enumerate(row_data):
        ws3.cell(row=r_idx, column=j + 1, value=val)
    style_data_row(ws3, r_idx, 7, alt=i % 2 == 0)

# Volume projection table
r = 12
add_subtitle(ws3, "  Proyeccion por Cantidad de Leads", r, 7, fill=PatternFill(start_color=ACCENT2, end_color=ACCENT2, fill_type="solid"))
r = 13
headers = ["Leads/mes", "Ventas estimadas", "Costo Claude", "Costo Voz", "Infraestructura", "TOTAL/mes", "Costo por Venta"]
for i, h in enumerate(headers):
    ws3.cell(row=r, column=i + 1, value=h)
style_header_row(ws3, r, 7)

volume_data = [
    [50,   "8-12",   5,  2,  10, 17,  1.70],
    [100,  "15-25",  10, 4,  10, 24,  1.20],
    [200,  "30-50",  20, 8,  12, 40,  1.00],
    [300,  "45-75",  30, 12, 12, 54,  0.90],
    [500,  "75-125", 50, 18, 15, 83,  0.83],
    [750,  "112-187",75, 26, 18, 119, 0.79],
    [1000, "150-250",100,35, 20, 155, 0.78],
]

for i, row_data in enumerate(volume_data):
    r_idx = r + 1 + i
    ws3.cell(row=r_idx, column=1, value=row_data[0])
    ws3.cell(row=r_idx, column=2, value=row_data[1])
    for j in range(2, 7):
        cell = ws3.cell(row=r_idx, column=j + 1, value=row_data[j])
        if j < 6:
            cell.number_format = '$#,##0'
        else:
            cell.number_format = '$#,##0.00'
    style_data_row(ws3, r_idx, 7, alt=i % 2 == 0)
    ws3.cell(row=r_idx, column=6).font = Font(name="Calibri", size=12, bold=True, color=ACCENT2)

# Chart: Monthly cost stacked bar
chart2 = BarChart()
chart2.type = "col"
chart2.title = "Costo Total Mensual por Volumen de Leads (USD)"
chart2.y_axis.title = "Costo USD/mes"
chart2.x_axis.title = "Cantidad de Leads"
chart2.style = 10
chart2.width = 28
chart2.height = 16

data_ref = Reference(ws3, min_col=3, max_col=5, min_row=13, max_row=20)
cats = Reference(ws3, min_col=1, min_row=14, max_row=20)
chart2.add_data(data_ref, titles_from_data=True)
chart2.set_categories(cats)
chart2.grouping = "stacked"

chart2.series[0].graphicalProperties.solidFill = BLUE
chart2.series[1].graphicalProperties.solidFill = ORANGE
chart2.series[2].graphicalProperties.solidFill = "95A5A6"

ws3.add_chart(chart2, "A22")

# Chart: Cost per sale line
chart3 = LineChart()
chart3.title = "Costo por Venta Cerrada vs Volumen"
chart3.y_axis.title = "USD por venta"
chart3.x_axis.title = "Leads/mes"
chart3.style = 10
chart3.width = 28
chart3.height = 16

data_ref = Reference(ws3, min_col=7, min_row=13, max_row=20)
cats = Reference(ws3, min_col=1, min_row=14, max_row=20)
chart3.add_data(data_ref, titles_from_data=True)
chart3.set_categories(cats)
chart3.series[0].graphicalProperties.line.solidFill = ACCENT
chart3.series[0].graphicalProperties.line.width = 28000
chart3.series[0].marker.symbol = "circle"
chart3.series[0].marker.size = 7

ws3.add_chart(chart3, "A39")

# ═══════════════════════════════════════════════════════════════
# SHEET 4: EFECTIVIDAD Y METRICAS
# ═══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Efectividad y Metricas")
ws4.sheet_properties.tabColor = PURPLE

for c in range(1, 7):
    ws4.column_dimensions[get_column_letter(c)].width = 22

add_title_banner(ws4, "METRICAS DE EFECTIVIDAD DEL BOT", 1, 6)

# Funnel
r = 3
add_subtitle(ws4, "  Embudo de Ventas (Lead Scoring)", r, 6, fill=PatternFill(start_color=ACCENT2, end_color=ACCENT2, fill_type="solid"))
r = 4
headers = ["Etapa", "Score", "% que llega", "Leads (de 100)", "Estrategia del Bot", "Accion"]
for i, h in enumerate(headers):
    ws4.cell(row=r, column=i + 1, value=h)
style_header_row(ws4, r, 6)

funnel_data = [
    ["NUEVO", "0-10", "100%", 100, "Bienvenida", "Saludo calido + pregunta"],
    ["ENGAGED", "11-30", "75%", 75, "Informar", "Responder dudas + beneficios"],
    ["INTERESADO", "31-60", "50%", 50, "Persuadir", "Social proof + urgencia"],
    ["CALIENTE", "61-80", "30%", 30, "Guiar", "Facilitar decision"],
    ["CIERRE", "81+", "20%", 20, "Cerrar", "Link de pago + confirmacion"],
]

colors_funnel = [BLUE, TEAL, YELLOW, ORANGE, GREEN]
for i, row_data in enumerate(funnel_data):
    r_idx = r + 1 + i
    for j, val in enumerate(row_data):
        ws4.cell(row=r_idx, column=j + 1, value=val)
    style_data_row(ws4, r_idx, 6, alt=i % 2 == 0)
    ws4.cell(row=r_idx, column=1).font = Font(name="Calibri", size=11, bold=True, color=colors_funnel[i])

# Funnel chart
chart_funnel = BarChart()
chart_funnel.type = "bar"
chart_funnel.title = "Embudo de Conversion (de cada 100 leads)"
chart_funnel.x_axis.title = "Cantidad de Leads"
chart_funnel.style = 10
chart_funnel.width = 28
chart_funnel.height = 14

data_ref = Reference(ws4, min_col=4, min_row=4, max_row=9)
cats = Reference(ws4, min_col=1, min_row=5, max_row=9)
chart_funnel.add_data(data_ref, titles_from_data=True)
chart_funnel.set_categories(cats)
chart_funnel.series[0].graphicalProperties.solidFill = ACCENT

ws4.add_chart(chart_funnel, "A11")

# Scoring signals
r = 27
add_subtitle(ws4, "  Senales que Detecta el Bot (Lead Scoring)", r, 6, fill=PatternFill(start_color=ACCENT2, end_color=ACCENT2, fill_type="solid"))
r = 28
headers = ["Senal del Cliente", "Puntos", "Ejemplo de Mensaje", "Intent Detectado", "Impacto", ""]
for i, h in enumerate(headers):
    ws4.cell(row=r, column=i + 1, value=h)
style_header_row(ws4, r, 6)

signals = [
    ["Saludo", "+10", "\"Hola, buenas!\"", "GREETING", "Bajo"],
    ["Pregunta de curso", "+15", "\"Que cursos tienen?\"", "COURSE_INQUIRY", "Medio"],
    ["Pregunta de precio", "+25", "\"Cuanto cuesta?\"", "PRICE_INQUIRY", "Alto"],
    ["Pregunta de horario", "+20", "\"Que dias hay clase?\"", "SCHEDULE_INQUIRY", "Alto"],
    ["Pregunta de ubicacion", "+17", "\"Donde queda la pista?\"", "LOCATION_INQUIRY", "Medio"],
    ["Senal de compra", "+35", "\"Me inscribo, como pago?\"", "BUY_SIGNAL", "Muy Alto"],
    ["Confirmacion de pago", "+40", "\"Ya transferi por Nequi\"", "PAYMENT_CONFIRM", "Critico"],
    ["Objecion de precio", "+15", "\"Esta muy caro\"", "OBJECTION_PRICE", "Medio"],
    ["Objecion de duda", "+15", "\"Es peligroso?\"", "OBJECTION_DOUBT", "Medio"],
    ["Precio + Horario combo", "+5 bonus", "Pregunto ambos", "COMBO", "Muy Alto"],
]

for i, row_data in enumerate(signals):
    r_idx = r + 1 + i
    for j, val in enumerate(row_data):
        ws4.cell(row=r_idx, column=j + 1, value=val)
    style_data_row(ws4, r_idx, 6, alt=i % 2 == 0)

    impact = row_data[4]
    impact_cell = ws4.cell(row=r_idx, column=5)
    if impact in ("Muy Alto", "Critico"):
        impact_cell.font = Font(name="Calibri", size=11, bold=True, color=GREEN)
    elif impact == "Alto":
        impact_cell.font = Font(name="Calibri", size=11, bold=True, color=ORANGE)

# KPIs
r = 40
add_subtitle(ws4, "  KPIs para Medir Efectividad", r, 6, fill=PatternFill(start_color=ACCENT2, end_color=ACCENT2, fill_type="solid"))
r = 41
headers = ["KPI", "Formula", "Meta", "Como se mide", "Frecuencia", ""]
for i, h in enumerate(headers):
    ws4.cell(row=r, column=i + 1, value=h)
style_header_row(ws4, r, 6)

kpis = [
    ["Tasa de Conversion", "Ventas / Leads totales", ">20%", "Score >= 81 / total leads", "Semanal"],
    ["Mensajes por Venta", "Total msgs ventas / ventas", "<15", "Promedio MessageCount", "Semanal"],
    ["Costo por Venta", "Gasto total / ventas", "<$1.50", "API costs / conversiones", "Mensual"],
    ["Tiempo de Respuesta", "Promedio latencia", "<3s texto, <6s audio", "Logs del servidor", "Diario"],
    ["Tasa de Abandono", "Leads sin responder 24h", "<40%", "Leads inactivos", "Semanal"],
    ["Satisfaccion", "Leads que agradecen", ">60%", "IntentThanks / total", "Mensual"],
    ["Objeciones Resueltas", "Objeciones -> Venta", ">50%", "Objection -> BuySignal", "Mensual"],
]

for i, row_data in enumerate(kpis):
    r_idx = r + 1 + i
    for j, val in enumerate(row_data):
        ws4.cell(row=r_idx, column=j + 1, value=val)
    style_data_row(ws4, r_idx, 6, alt=i % 2 == 0)
    ws4.cell(row=r_idx, column=1).font = body_bold

# ═══════════════════════════════════════════════════════════════
# SHEET 5: BOT vs HUMANO
# ═══════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Bot vs Humano")
ws5.sheet_properties.tabColor = TEAL

for c in range(1, 6):
    ws5.column_dimensions[get_column_letter(c)].width = 24

add_title_banner(ws5, "COMPARATIVA: BOT vs ASESOR HUMANO", 1, 5)

r = 3
headers = ["Criterio", "Asesor Humano", "Bot IA", "Ahorro/Ventaja", "Ganador"]
for i, h in enumerate(headers):
    ws5.cell(row=r, column=i + 1, value=h)
style_header_row(ws5, r, 5)

comparison = [
    ["Costo mensual (100 leads)", "$400-800 USD (salario)", "$24 USD", "94-97% ahorro", "BOT"],
    ["Costo mensual (500 leads)", "$1,200+ USD (2-3 asesores)", "$83 USD", "93% ahorro", "BOT"],
    ["Disponibilidad", "8-10 horas/dia", "24/7/365", "14-16 hrs extra/dia", "BOT"],
    ["Tiempo de respuesta", "5-30 minutos", "2-5 segundos", "99% mas rapido", "BOT"],
    ["Consistencia", "Variable (humor, cansancio)", "100% consistente", "Sin dias malos", "BOT"],
    ["Escalabilidad", "Lineal (mas gente)", "Casi infinita", "Sin contrataciones", "BOT"],
    ["Tono natural", "Nativo, autentico", "Muy bueno (entrenado)", "Similar", "EMPATE"],
    ["Objeciones complejas", "Excelente (experiencia)", "Bueno (con prompts)", "Humano mejor", "HUMANO"],
    ["Empatia genuina", "Alta", "Simulada (convincente)", "Humano mejor", "HUMANO"],
    ["Voz personalizada", "Natural", "Clonada (ElevenLabs)", "Practicamente igual", "EMPATE"],
    ["Seguimiento automatico", "Se olvida", "Programable", "Bot no se olvida", "BOT"],
    ["Reportes y metricas", "Manual", "Automatico (scoring)", "Datos tiempo real", "BOT"],
]

for i, row_data in enumerate(comparison):
    r_idx = r + 1 + i
    for j, val in enumerate(row_data):
        ws5.cell(row=r_idx, column=j + 1, value=val)
    style_data_row(ws5, r_idx, 5, alt=i % 2 == 0)

    winner = row_data[4]
    winner_cell = ws5.cell(row=r_idx, column=5)
    if winner == "BOT":
        winner_cell.font = Font(name="Calibri", size=11, bold=True, color=GREEN)
        winner_cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    elif winner == "HUMANO":
        winner_cell.font = Font(name="Calibri", size=11, bold=True, color=ORANGE)
        winner_cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    else:
        winner_cell.font = Font(name="Calibri", size=11, bold=True, color=BLUE)
        winner_cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")

# Pie chart data
r_pie = 17
ws5.cell(row=r_pie, column=1, value="Resultado").font = header_font
ws5.cell(row=r_pie, column=1).fill = header_fill
ws5.cell(row=r_pie, column=2, value="Cantidad").font = header_font
ws5.cell(row=r_pie, column=2).fill = header_fill
ws5.cell(row=r_pie + 1, column=1, value="Bot gana")
ws5.cell(row=r_pie + 1, column=2, value=8)
ws5.cell(row=r_pie + 2, column=1, value="Humano gana")
ws5.cell(row=r_pie + 2, column=2, value=2)
ws5.cell(row=r_pie + 3, column=1, value="Empate")
ws5.cell(row=r_pie + 3, column=2, value=2)

chart_pie = PieChart()
chart_pie.title = "Bot vs Humano - Resultado por Criterio"
chart_pie.style = 10
chart_pie.width = 18
chart_pie.height = 14

data_ref = Reference(ws5, min_col=2, min_row=r_pie, max_row=r_pie + 3)
cats = Reference(ws5, min_col=1, min_row=r_pie + 1, max_row=r_pie + 3)
chart_pie.add_data(data_ref, titles_from_data=True)
chart_pie.set_categories(cats)

pt0 = DataPoint(idx=0)
pt0.graphicalProperties.solidFill = GREEN
pt1 = DataPoint(idx=1)
pt1.graphicalProperties.solidFill = ORANGE
pt2 = DataPoint(idx=2)
pt2.graphicalProperties.solidFill = BLUE
chart_pie.series[0].data_points = [pt0, pt1, pt2]

chart_pie.dataLabels = DataLabelList()
chart_pie.dataLabels.showPercent = True
chart_pie.dataLabels.showVal = True

ws5.add_chart(chart_pie, "A22")

# ROI
r = 39
add_subtitle(ws5, "  Retorno de Inversion (ROI)", r, 5, fill=PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid"))

roi_headers = ["Escenario", "Leads/mes", "Costo Bot", "Costo Humano", "Ahorro Mensual"]
r2 = r + 1
for i, h in enumerate(roi_headers):
    ws5.cell(row=r2, column=i + 1, value=h)
style_header_row(ws5, r2, 5)

roi_data = [
    ["Pequeno", "50", "$17/mes", "$400/mes", "$383/mes"],
    ["Mediano", "200", "$40/mes", "$800/mes", "$760/mes"],
    ["Grande", "500", "$83/mes", "$1,500/mes", "$1,417/mes"],
    ["Enterprise", "1,000", "$155/mes", "$3,000/mes", "$2,845/mes"],
]

for i, row_data in enumerate(roi_data):
    r_idx = r2 + 1 + i
    for j, val in enumerate(row_data):
        ws5.cell(row=r_idx, column=j + 1, value=val)
    style_data_row(ws5, r_idx, 5, alt=i % 2 == 0)
    ws5.cell(row=r_idx, column=5).font = Font(name="Calibri", size=12, bold=True, color=GREEN)

# ═══════════════════════════════════════════════════════════════
# SHEET 6: RECOMENDACIONES
# ═══════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Recomendaciones")
ws6.sheet_properties.tabColor = YELLOW

for c in range(1, 5):
    ws6.column_dimensions[get_column_letter(c)].width = 30

add_title_banner(ws6, "RECOMENDACIONES Y PROXIMOS PASOS", 1, 4)

r = 3
add_subtitle(ws6, "  Optimizaciones para Reducir Costos", r, 4, fill=PatternFill(start_color=ACCENT2, end_color=ACCENT2, fill_type="solid"))

r = 4
headers = ["Optimizacion", "Ahorro Estimado", "Impacto", "Prioridad"]
for i, h in enumerate(headers):
    ws6.cell(row=r, column=i + 1, value=h)
style_header_row(ws6, r, 4)

opts = [
    ["Cache de respuestas frecuentes (Redis)", "20-30%", "Alto - muchas preguntas se repiten", "ALTA"],
    ["Haiku para intent classification", "No aplica (ya es local)", "N/A - ya optimizado", "HECHO"],
    ["Comprimir historial > 10 mensajes", "15-20%", "Medio - reduce tokens input", "MEDIA"],
    ["Cache de audio TTS frases comunes", "30-40% en voz", "Alto - saludos y cierre repiten", "ALTA"],
    ["Delay aleatorio 2-8s en respuesta", "$0 (gratis)", "Alto - se siente mas humano", "ALTA"],
    ["Persistir Lead Scores en SQLite", "$0 (gratis)", "Critico - metricas reales", "CRITICA"],
    ["Dashboard de metricas", "$0 (gratis)", "Alto - visibilidad para cliente", "ALTA"],
]

for i, row_data in enumerate(opts):
    r_idx = r + 1 + i
    for j, val in enumerate(row_data):
        ws6.cell(row=r_idx, column=j + 1, value=val)
    style_data_row(ws6, r_idx, 4, alt=i % 2 == 0)

    prio = row_data[3]
    prio_cell = ws6.cell(row=r_idx, column=4)
    if prio in ("ALTA", "CRITICA"):
        prio_cell.font = Font(name="Calibri", size=11, bold=True, color=RED)
    elif prio == "MEDIA":
        prio_cell.font = Font(name="Calibri", size=11, bold=True, color=ORANGE)
    else:
        prio_cell.font = Font(name="Calibri", size=11, bold=True, color=GREEN)

# Humanization
r = 13
add_subtitle(ws6, "  Estrategia de Humanizacion", r, 4, fill=PatternFill(start_color=ACCENT2, end_color=ACCENT2, fill_type="solid"))

r = 14
headers = ["Estrategia", "Como Funciona", "Estado Actual", "Impacto en Ventas"]
for i, h in enumerate(headers):
    ws6.cell(row=r, column=i + 1, value=h)
style_header_row(ws6, r, 4)

human_data = [
    ["Tono colombiano natural", "Usa 'parce', 'de una', 'dale' calibrado", "IMPLEMENTADO", "Alto - genera confianza"],
    ["Mensajes cortos (max 2 oraciones)", "max_tokens=300, regla en prompt", "IMPLEMENTADO", "Alto - nadie lee parrafos"],
    ["Voz clonada del dueno", "ElevenLabs con muestras reales", "IMPLEMENTADO", "Muy Alto - la gente confia"],
    ["Siempre termina con pregunta", "Regla en system prompt", "IMPLEMENTADO", "Critico - mantiene conversacion"],
    ["Delay de respuesta aleatorio", "Esperar 2-8s antes de responder", "PENDIENTE", "Alto - elimina sensacion bot"],
    ["Aprender de ventas reales", "Google Sheets con ejemplos", "PARCIAL", "Muy Alto - imita vendedor"],
    ["Variacion de expresiones", "Pool de frases por contexto", "PARCIAL", "Medio - evita repeticion"],
    ["Deteccion de humor/emocion", "Ajustar tono segun estado lead", "PENDIENTE", "Alto - empatia real"],
]

for i, row_data in enumerate(human_data):
    r_idx = r + 1 + i
    for j, val in enumerate(row_data):
        ws6.cell(row=r_idx, column=j + 1, value=val)
    style_data_row(ws6, r_idx, 4, alt=i % 2 == 0)

    estado = row_data[2]
    estado_cell = ws6.cell(row=r_idx, column=3)
    if estado == "IMPLEMENTADO":
        estado_cell.font = Font(name="Calibri", size=11, bold=True, color=GREEN)
    elif estado == "PARCIAL":
        estado_cell.font = Font(name="Calibri", size=11, bold=True, color=ORANGE)
    else:
        estado_cell.font = Font(name="Calibri", size=11, bold=True, color=RED)

# Footer
r = 24
ws6.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
ws6.cell(row=r, column=1,
         value="Nota: Costos estimados basados en precios publicos de Anthropic y ElevenLabs a abril 2026. Los costos reales pueden variar.").font = small_font

# ── Save ──
filepath = r"D:\Informe_Costos_Bot_Instagram_StageBot.xlsx"
wb.save(filepath)
print(f"Informe guardado en: {filepath}")
